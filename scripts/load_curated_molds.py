"""Q.54.R — backfill do master de moldes na camada curada.

A folha ``Moldes`` de ``Folha_IA_extra.xlsx`` (510 moldes) nunca foi
carregada para ``factory_curated.mold`` — só as tabelas-irmãs do mesmo
Excel (``quality_event``, ``order_phase``, ``allocation``). Sem ela, os
252 moldes reais que aparecem nos dados de defeitos não têm nome nem
tipo, e a área de Moldes do software fica desligada de `plan.mold`
(catálogo sintético, espaço de IDs diferente).

Este script carrega o master em falta, ligado à **mesma ingestão** das
tabelas-irmãs (mesma proveniência — `ingestion_id`). O mapeamento é
idêntico ao de :meth:`IngestTransformer._transform_molds`.

Idempotente: apaga as linhas de mold dessa ingestão e volta a inserir.
Isolado: `factory_curated.mold` não tem nada a apontar para ela (os
`molde_id` nas outras tabelas são strings, não FKs), por isso carregar
não pode partir a camada curada.

Uso:
    .venv/Scripts/python.exe scripts/load_curated_molds.py
"""

from __future__ import annotations

import asyncio
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select

from src.factory_data_product.models.curated import (
    CuratedMold,
    CuratedOrderPhase,
)
from src.shared.database import async_session_factory

WORKBOOK = Path(__file__).resolve().parent.parent / "Folha_IA_extra.xlsx"
SHEET = "Moldes"


def _s(value: object) -> str | None:
    """Equivalente ao `_safe_str` do transformer: str trimmed, None se vazio."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _read_moldes() -> list[dict[str, object]]:
    """Lê a folha Moldes → lista de dicts {coluna: valor}."""
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Folha '{SHEET}' não existe em {WORKBOOK.name}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, object]] = []
    for raw in rows[1:]:
        if raw is None or all(c is None for c in raw):
            continue
        out.append(dict(zip(header, raw)))
    return out


def _to_mold(payload: dict[str, object], ingestion_id) -> CuratedMold | None:
    """Mapeia uma linha Moldes → CuratedMold (mesmo mapa do transformer)."""
    molde_id = _s(payload.get("MoldeId"))
    if not molde_id:
        return None
    estado = _s(payload.get("MoldeEstado"))
    # NELO usa estado=4 para "manutenção" (mesma derivação do transformer).
    em_manutencao = estado == "4"
    return CuratedMold(
        ingestion_id=ingestion_id,
        molde_id=molde_id,
        molde_nome=_s(payload.get("MoldeNome")),
        modelo_id=_s(payload.get("MoldeModeloId")),
        tamanho_id=_s(payload.get("MoldeTamanhoId")),
        # MoldeModelo é a string descritiva ("K1 Vanquish II ML") — o
        # transformer usa-a como `tipo` por falta de campo dedicado.
        tipo=_s(payload.get("MoldeModelo")),
        estado=estado,
        em_manutencao=em_manutencao,
    )


async def main() -> int:
    moldes = _read_moldes()
    if not moldes:
        print("Folha Moldes vazia — nada a carregar.")
        return 1

    async with async_session_factory() as session:
        # Proveniência: usar a ingestão da camada curada já existente.
        ingestion_id = (
            await session.execute(select(CuratedOrderPhase.ingestion_id).limit(1))
        ).scalar_one_or_none()
        if ingestion_id is None:
            print(
                "factory_curated.order_phase está vazio — corre primeiro o "
                "ingest da camada curada. Abortado para não inventar "
                "uma ingestão órfã."
            )
            return 1

        # Idempotência: limpar moldes desta ingestão antes de reinserir.
        await session.execute(
            delete(CuratedMold).where(CuratedMold.ingestion_id == ingestion_id)
        )

        rows = [m for m in (_to_mold(p, ingestion_id) for p in moldes) if m]
        session.add_all(rows)
        await session.commit()

    print(
        f"factory_curated.mold carregado: {len(rows)} moldes "
        f"(ingestão {ingestion_id})."
    )
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
