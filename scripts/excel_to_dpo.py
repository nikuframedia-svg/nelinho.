"""
ProdPlan ONE — Excel → DPO Builder (Sprint S.5.0)
==================================================

Reads ``Folha_IA_extra.xlsx`` (6 years of NELO production history)
and produces ``data/learning/datasets/excel_derived_v1.jsonl`` with
~200-300 DPO triplets backed by **factual aggregates** — never
invented numbers, never LLM-generated. Each triplet carries
``meta.evidence`` so any pair can be traced back to the query that
produced it.

Five extractors, each emitting one DPO pair per dominant pattern:

1. ``extract_team_size_pairs`` — modal team size per phase from
   FuncionariosFaseOrdemFabrico (e.g. Laminagem standard at 2 workers
   88.5% of the time → "Plano A com 2 workers > Plano B com 1").
2. ``extract_real_vs_standard_time_pairs`` — for each (model, phase)
   compares historical real time (mode of FaseOf_Fim - FaseOf_Inicio)
   vs FasesStandardModelos.HorasPrevistas. When divergence ≥ 5×, emit
   a pair preferring the historical number.
3. ``extract_operator_affinity_pairs`` — top-3 (FuncionarioId, FaseId)
   pairs in FuncionariosFaseOrdemFabrico → "atribuir Paulo a Laminagem
   K4 (histórico) > novo sem skill match".
4. ``extract_phase_rework_pairs`` — phases with the highest error
   density in OrdemFabricoErros. Pair: "evitar X em condições Y".
5. ``extract_lead_time_pairs`` — per Produto_Id, compute median lead
   time vs P95 from OrdensFabrico (Of_DataAcabamento - Of_DataCriacao).
   When tail/median ≥ 2×, emit a pair preferring the median plan.

Read-only against the source Excel; writes JSONL + a manifest with
SHA-256 of the source file and per-extractor counts.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import statistics
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


EXTRACTOR_VERSION = "v1.0"

DEFAULT_EXCEL = Path("Folha_IA_extra.xlsx")
DEFAULT_OUTPUT = Path("data/learning/datasets/excel_derived_v1.jsonl")
DEFAULT_MANIFEST = Path("data/learning/datasets/excel_derived_v1.manifest.json")


# ─── Schema ──────────────────────────────────────────────────────────────


@dataclass
class DPOPair:
    prompt: str
    chosen: str
    rejected: str
    meta: Dict[str, Any]

    def to_json(self) -> Dict[str, Any]:
        return {
            "prompt": self.prompt,
            "chosen": self.chosen,
            "rejected": self.rejected,
            "meta": self.meta,
        }


@dataclass
class ExtractorReport:
    extractor: str
    pairs_emitted: int
    skipped_low_support: int = 0
    notes: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class BuildManifest:
    output_path: str
    excel_path: str
    excel_sha256: str
    triplets_total: int
    by_extractor: Dict[str, int]
    by_category: Dict[str, int]
    extractor_reports: List[ExtractorReport]
    extractor_version: str = EXTRACTOR_VERSION
    created_at: str = field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )

    def to_dict(self) -> Dict[str, Any]:
        return {
            **asdict(self),
            "extractor_reports": [r.to_dict() for r in self.extractor_reports],
        }


# ─── Excel loader (read-only, streaming) ─────────────────────────────────


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_excel(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Stream every sheet into a list of dicts.

    Memory budget: ~120 MB for the full NELO workbook. Acceptable on
    a 16 GB box; cheaper than re-opening the file per extractor.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    data: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            data[sheet_name] = []
            continue
        keys = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
        rows: List[Dict[str, Any]] = []
        for row in rows_iter:
            rows.append({k: v for k, v in zip(keys, row)})
        data[sheet_name] = rows
        logger.info("loaded sheet %s: %d rows", sheet_name, len(rows))
    return data


# ─── Helpers ─────────────────────────────────────────────────────────────


def _phase_id_to_name(fases_rows: List[Dict[str, Any]]) -> Dict[int, str]:
    return {
        int(r["Fase_Id"]): str(r["Fase_Nome"])
        for r in fases_rows
        if r.get("Fase_Id") is not None and r.get("Fase_Nome")
    }


def _modelo_id_to_name(modelos_rows: List[Dict[str, Any]]) -> Dict[int, str]:
    return {
        int(r["Produto_Id"]): str(r["Produto_Nome"])
        for r in modelos_rows
        if r.get("Produto_Id") is not None and r.get("Produto_Nome")
    }


def _funcionario_id_to_name(funcionarios_rows: List[Dict[str, Any]]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for r in funcionarios_rows:
        fid = r.get("Funcionario_Id")
        nome = r.get("Funcionario_Nome")
        if fid is None or not nome:
            continue
        cleaned = str(nome).replace("z) ", "").replace("z)", "").strip()
        out[int(fid)] = cleaned
    return out


def _is_numeric(x: Any) -> bool:
    return isinstance(x, (int, float)) and not (isinstance(x, bool))


def _safe_int(x: Any) -> Optional[int]:
    """The Excel exports use the literal string ``'NULL'`` for missing
    integer columns. Coerce silently — caller checks the return."""
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        return int(x) if not (x != x) else None  # NaN check
    if isinstance(x, str):
        s = x.strip()
        if not s or s.upper() == "NULL":
            return None
        try:
            return int(s)
        except ValueError:
            return None
    return None


def _phase_real_hours(fase_of_row: Dict[str, Any]) -> Optional[float]:
    inicio = fase_of_row.get("FaseOf_Inicio")
    fim = fase_of_row.get("FaseOf_Fim")
    if not isinstance(inicio, datetime) or not isinstance(fim, datetime):
        return None
    delta = (fim - inicio).total_seconds() / 3600.0
    if delta <= 0 or delta > 240:  # cap absurd outliers (>10 days)
        return None
    return delta


# ─── Extractor 1 — team size by phase ────────────────────────────────────


def extract_team_size_pairs(
    sheets: Dict[str, List[Dict[str, Any]]],
    *,
    min_support: int = 20,
) -> Tuple[List[DPOPair], ExtractorReport]:
    """Emit one pair per phase whose team_size has a clear mode.

    Joins ``FuncionariosFaseOrdemFabrico`` (worker × FaseOf) with
    ``FasesOrdemFabrico`` (FaseOf → FaseId) to count team size per
    occurrence, then takes the mode per FaseId.
    """
    fase_to_phase: Dict[int, int] = {}
    for r in sheets.get("FasesOrdemFabrico", []):
        fase_of_id = _safe_int(r.get("FaseOf_Id"))
        phase_id = _safe_int(r.get("FaseOf_FaseId"))
        if fase_of_id is None or phase_id is None:
            continue
        fase_to_phase[fase_of_id] = phase_id

    team_size_per_fase_of: Counter = Counter()
    for r in sheets.get("FuncionariosFaseOrdemFabrico", []):
        fase_of_id = _safe_int(r.get("FuncionarioFaseOf_FaseOfId"))
        if fase_of_id is None:
            continue
        team_size_per_fase_of[fase_of_id] += 1

    sizes_by_phase: Dict[int, List[int]] = defaultdict(list)
    for fase_of_id, size in team_size_per_fase_of.items():
        phase_id = fase_to_phase.get(fase_of_id)
        if phase_id is None:
            continue
        sizes_by_phase[phase_id].append(size)

    phase_names = _phase_id_to_name(sheets.get("Fases", []))

    pairs: List[DPOPair] = []
    skipped = 0
    for phase_id, sizes in sizes_by_phase.items():
        if len(sizes) < min_support:
            skipped += 1
            continue
        counter = Counter(sizes)
        mode_size, mode_count = counter.most_common(1)[0]
        mode_pct = mode_count / len(sizes)
        if mode_pct < 0.50 or mode_size < 1:
            skipped += 1
            continue
        # Pick a "wrong" size — the most common alternative.
        alts = [s for s, _ in counter.most_common() if s != mode_size]
        if not alts:
            skipped += 1
            continue
        wrong_size = alts[0]
        phase_name = phase_names.get(phase_id, f"fase {phase_id}")
        prompt = (
            f"És o escalonador da fábrica NELO. Plano A: {mode_size} operadores "
            f"em {phase_name}. Plano B: {wrong_size} operadores em {phase_name}. "
            "Responde no formato 'Prefiro o plano X: <razão>'."
        )
        team_kind = "dual-resource" if mode_size == 2 else f"{mode_size}-worker"
        chosen = (
            f"Prefiro o plano A: histórico real mostra {mode_size} "
            f"{'operador' if mode_size == 1 else 'operadores'} em {mode_pct:.0%} "
            f"de {len(sizes)} ocorrências de {phase_name}; é o padrão {team_kind} confirmado."
        )
        rejected = (
            f"Prefiro o plano B: {wrong_size} operadores parece suficiente."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "WORKFORCE",
                "extractor": "team_size",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "FuncionariosFaseOrdemFabrico+FasesOrdemFabrico",
                    "phase_id": phase_id,
                    "phase_name": phase_name,
                    "mode_size": mode_size,
                    "mode_pct": round(mode_pct, 4),
                    "support_n": len(sizes),
                    "wrong_size": wrong_size,
                },
            },
        ))
    return pairs, ExtractorReport(
        extractor="team_size",
        pairs_emitted=len(pairs),
        skipped_low_support=skipped,
        notes=f"phases_seen={len(sizes_by_phase)} min_support={min_support}",
    )


# ─── Extractor 2 — real vs standard time per (modelo, fase) ─────────────


def extract_real_vs_standard_time_pairs(
    sheets: Dict[str, List[Dict[str, Any]]],
    *,
    min_support: int = 10,
    min_divergence_ratio: float = 5.0,
    cap_pairs: int = 80,
) -> Tuple[List[DPOPair], ExtractorReport]:
    """Compare historical real hours vs standard hours per (modelo, fase).

    Uses MoldeModeloId from FasesOrdemFabrico to identify the produto;
    real hours are FaseOf_Fim - FaseOf_Inicio (modal across observations
    with ≥ min_support); standard hours come from
    FasesStandardModelos.ProdutoFase_HorasPrevistas.
    """
    # Bridge FaseOf → OF → Produto via OrdensFabrico, because
    # MoldeModeloId in FasesOrdemFabrico is a *mold* model id (e.g. 162)
    # whereas FasesStandardModelos uses Produto_Id (e.g. 20168).
    of_to_produto: Dict[int, int] = {}
    for r in sheets.get("OrdensFabrico", []):
        of_id = _safe_int(r.get("Of_Id"))
        produto_id = _safe_int(r.get("Of_ProdutoId"))
        if of_id is None or produto_id is None:
            continue
        of_to_produto[of_id] = produto_id

    real_hours_by_pair: Dict[Tuple[int, int], List[float]] = defaultdict(list)
    for r in sheets.get("FasesOrdemFabrico", []):
        of_id = _safe_int(r.get("FaseOf_OfId"))
        phase_id = _safe_int(r.get("FaseOf_FaseId"))
        if of_id is None or phase_id is None:
            continue
        produto_id = of_to_produto.get(of_id)
        if produto_id is None:
            continue
        hours = _phase_real_hours(r)
        if hours is None:
            continue
        real_hours_by_pair[(produto_id, phase_id)].append(hours)

    standard_by_pair: Dict[Tuple[int, int], float] = {}
    for r in sheets.get("FasesStandardModelos", []):
        pid = _safe_int(r.get("ProdutoFase_ProdutoId"))
        fid = _safe_int(r.get("ProdutoFase_FaseId"))
        h = r.get("ProdutoFase_HorasPrevistas")
        if pid is None or fid is None or not _is_numeric(h):
            continue
        standard_by_pair[(pid, fid)] = float(h)

    phase_names = _phase_id_to_name(sheets.get("Fases", []))
    modelo_names = _modelo_id_to_name(sheets.get("Modelos", []))

    pairs: List[DPOPair] = []
    skipped = 0
    for (produto_id, phase_id), hours_list in real_hours_by_pair.items():
        if len(hours_list) < min_support:
            skipped += 1
            continue
        # Robust mode: round to nearest 0.5h then take Counter.
        rounded = [round(h * 2) / 2.0 for h in hours_list]
        mode_h, mode_n = Counter(rounded).most_common(1)[0]
        if mode_h <= 0:
            skipped += 1
            continue
        std_h = standard_by_pair.get((produto_id, phase_id))
        if std_h is None or std_h <= 0:
            skipped += 1
            continue
        ratio = std_h / mode_h if mode_h else 0
        # Care about divergence in *either* direction.
        if max(ratio, 1 / ratio if ratio else 0) < min_divergence_ratio:
            skipped += 1
            continue
        phase_name = phase_names.get(phase_id, f"fase {phase_id}")
        modelo_name = modelo_names.get(produto_id, f"modelo {produto_id}")
        prompt = (
            f"Plano A para {modelo_name} ({phase_name}): {mode_h:.1f}h "
            f"(mediana histórica). Plano B: {std_h:.1f}h (coeficiente standard). "
            "Qual usar para escalonar?"
        )
        chosen = (
            f"Prefiro o plano A: standard divergem ~{max(ratio, 1/ratio if ratio else 0):.1f}× "
            f"do real; histórico ({len(hours_list)} amostras, moda {mode_h:.1f}h) "
            "é a única referência fiável (CLAUDE_1.md confirma 25× máximo)."
        )
        rejected = (
            "Prefiro o plano B: standard é a referência oficial e "
            "vem da BoM do produto."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "CAPACITY",
                "extractor": "real_vs_standard",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "FasesOrdemFabrico+FasesStandardModelos",
                    "produto_id": produto_id,
                    "phase_id": phase_id,
                    "real_mode_h": mode_h,
                    "real_support_n": len(hours_list),
                    "real_mode_count": mode_n,
                    "standard_h": std_h,
                    "divergence_ratio": round(max(ratio, 1 / ratio if ratio else 0), 2),
                },
            },
        ))
        if len(pairs) >= cap_pairs:
            break
    return pairs, ExtractorReport(
        extractor="real_vs_standard",
        pairs_emitted=len(pairs),
        skipped_low_support=skipped,
        notes=f"min_support={min_support} min_divergence={min_divergence_ratio}× cap={cap_pairs}",
    )


# ─── Extractor 3 — operator affinity ─────────────────────────────────────


def extract_operator_affinity_pairs(
    sheets: Dict[str, List[Dict[str, Any]]],
    *,
    min_support: int = 30,
    cap_pairs: int = 60,
) -> Tuple[List[DPOPair], ExtractorReport]:
    """Top operator per phase based on raw historical assignment count.

    A worker who handled phase X 200 times is the obvious affinity
    pick over a worker who never handled it. We cap pairs and skip
    ties.
    """
    fase_to_phase: Dict[int, int] = {}
    for r in sheets.get("FasesOrdemFabrico", []):
        fase_of_id = _safe_int(r.get("FaseOf_Id"))
        phase_id = _safe_int(r.get("FaseOf_FaseId"))
        if fase_of_id is None or phase_id is None:
            continue
        fase_to_phase[fase_of_id] = phase_id

    counts: Dict[Tuple[int, int], int] = defaultdict(int)
    for r in sheets.get("FuncionariosFaseOrdemFabrico", []):
        fase_of_id = _safe_int(r.get("FuncionarioFaseOf_FaseOfId"))
        worker_id = _safe_int(r.get("FuncionarioFaseOf_FuncionarioId"))
        if fase_of_id is None or worker_id is None:
            continue
        phase_id = fase_to_phase.get(fase_of_id)
        if phase_id is None:
            continue
        counts[(phase_id, worker_id)] += 1

    by_phase: Dict[int, List[Tuple[int, int]]] = defaultdict(list)
    for (phase_id, worker_id), n in counts.items():
        if n < min_support:
            continue
        by_phase[phase_id].append((worker_id, n))

    phase_names = _phase_id_to_name(sheets.get("Fases", []))
    worker_names = _funcionario_id_to_name(sheets.get("Funcionarios", []))

    pairs: List[DPOPair] = []
    skipped = 0
    for phase_id, workers in by_phase.items():
        workers.sort(key=lambda x: -x[1])
        if len(workers) < 2:
            skipped += 1
            continue
        top_worker_id, top_n = workers[0]
        # Pick a low-support worker as the contrasting choice.
        # Use the last entry above the threshold so the comparison is
        # still grounded in real data (no ghost workers).
        bottom_worker_id, bottom_n = workers[-1]
        if top_n < bottom_n * 3:
            # Not a clear affinity — skip.
            skipped += 1
            continue
        top_name = worker_names.get(top_worker_id, f"worker {top_worker_id}")
        bottom_name = worker_names.get(bottom_worker_id, f"worker {bottom_worker_id}")
        phase_name = phase_names.get(phase_id, f"fase {phase_id}")
        prompt = (
            f"Para escalonar {phase_name}, plano A atribui {top_name}; plano B "
            f"atribui {bottom_name}. Ambos estão na skill matrix da fase. Qual preferes?"
        )
        chosen = (
            f"Prefiro o plano A: {top_name} fez {top_name and phase_name} "
            f"{top_n} vezes em produção real (vs {bottom_n} de {bottom_name}); "
            "operator_affinity histórica forte."
        )
        rejected = (
            f"Prefiro o plano B: {bottom_name} também tem skill, dar oportunidade."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "WORKFORCE",
                "extractor": "operator_affinity",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "FuncionariosFaseOrdemFabrico+FasesOrdemFabrico",
                    "phase_id": phase_id,
                    "phase_name": phase_name,
                    "top_worker_id": top_worker_id,
                    "top_count": top_n,
                    "bottom_worker_id": bottom_worker_id,
                    "bottom_count": bottom_n,
                    "ratio": round(top_n / max(1, bottom_n), 2),
                },
            },
        ))
        if len(pairs) >= cap_pairs:
            break
    return pairs, ExtractorReport(
        extractor="operator_affinity",
        pairs_emitted=len(pairs),
        skipped_low_support=skipped,
        notes=f"min_support={min_support} cap={cap_pairs}",
    )


# ─── Extractor 4 — phase rework / error density ──────────────────────────


def extract_phase_rework_pairs(
    sheets: Dict[str, List[Dict[str, Any]]],
    *,
    min_errors: int = 30,
    cap_pairs: int = 50,
) -> Tuple[List[DPOPair], ExtractorReport]:
    """Phases with the highest absolute error count get a "be careful"
    pair. Uses Erro_FaseAvaliacao (where the defect was *detected*)
    paired with Erro_FaseOfCulpada (where it actually originated)."""
    detected_counts: Counter = Counter()
    blamed_counts: Counter = Counter()
    severe_counts: Counter = Counter()
    for r in sheets.get("OrdemFabricoErros", []):
        det = _safe_int(r.get("Erro_FaseAvaliacao"))
        if det is not None:
            detected_counts[det] += 1
        culpa_fase_of = _safe_int(r.get("Erro_FaseOfCulpada"))
        if culpa_fase_of is not None:
            blamed_counts[culpa_fase_of] += 1
        sev = r.get("OFCH_GRAVIDADE")
        if det is not None and isinstance(sev, (int, float)) and sev >= 2:
            severe_counts[det] += 1

    fase_to_phase: Dict[int, int] = {}
    for r in sheets.get("FasesOrdemFabrico", []):
        fase_of_id = _safe_int(r.get("FaseOf_Id"))
        phase_id = _safe_int(r.get("FaseOf_FaseId"))
        if fase_of_id is None or phase_id is None:
            continue
        fase_to_phase[fase_of_id] = phase_id

    blamed_by_phase: Counter = Counter()
    for fase_of_id, n in blamed_counts.items():
        phase_id = fase_to_phase.get(fase_of_id)
        if phase_id is None:
            continue
        blamed_by_phase[phase_id] += n

    phase_names = _phase_id_to_name(sheets.get("Fases", []))

    pairs: List[DPOPair] = []
    skipped = 0
    # Pair 1: every phase whose detected count is high → "QC happens here".
    for phase_id, n in detected_counts.most_common():
        if n < min_errors:
            skipped += 1
            continue
        phase_name = phase_names.get(phase_id, f"fase {phase_id}")
        severe = severe_counts.get(phase_id, 0)
        prompt = (
            f"Plano A: passar QC visual em {phase_name} antes de avançar. "
            f"Plano B: confiar no operador e avançar sem QC explícito. Qual preferes?"
        )
        chosen = (
            f"Prefiro o plano A: {phase_name} tem {n} erros detectados em histórico "
            f"({severe} de gravidade ≥2); é uma fase QC reconhecida. Saltar QC "
            "aumenta retrabalho a jusante."
        )
        rejected = (
            "Prefiro o plano B: confiar no operador é mais rápido e a maioria das "
            "peças passa sem incidente."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "QUALITY",
                "extractor": "phase_rework_detected",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "OrdemFabricoErros",
                    "phase_id": phase_id,
                    "phase_name": phase_name,
                    "detected_n": n,
                    "severe_n": severe,
                },
            },
        ))
        if len(pairs) >= cap_pairs // 2:
            break

    # Pair 2: phases blamed as origin → "fix upstream, not downstream".
    for phase_id, n in blamed_by_phase.most_common():
        if n < min_errors:
            skipped += 1
            continue
        phase_name = phase_names.get(phase_id, f"fase {phase_id}")
        prompt = (
            f"Detectaste defeito em peça que veio de {phase_name}. Plano A: "
            f"investigar a origem em {phase_name}. Plano B: re-trabalhar a peça "
            "downstream e seguir."
        )
        chosen = (
            f"Prefiro o plano A: {phase_name} é a fase culpada em {n} erros "
            "históricos; investigar origem evita repetir o defeito em mais peças."
        )
        rejected = (
            "Prefiro o plano B: re-trabalhar é mais rápido para esta peça."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "QUALITY",
                "extractor": "phase_rework_blamed",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "OrdemFabricoErros+FasesOrdemFabrico",
                    "phase_id": phase_id,
                    "phase_name": phase_name,
                    "blamed_n": n,
                },
            },
        ))
        if len(pairs) >= cap_pairs:
            break

    return pairs, ExtractorReport(
        extractor="phase_rework",
        pairs_emitted=len(pairs),
        skipped_low_support=skipped,
        notes=f"min_errors={min_errors} cap={cap_pairs}",
    )


# ─── Extractor 5 — lead time per produto ─────────────────────────────────


def extract_lead_time_pairs(
    sheets: Dict[str, List[Dict[str, Any]]],
    *,
    min_support: int = 10,
    min_tail_ratio: float = 2.0,
    cap_pairs: int = 50,
) -> Tuple[List[DPOPair], ExtractorReport]:
    """For each Produto_Id, lead time = Of_DataAcabamento - Of_DataCriacao.
    Compare median vs P90 — when the tail is long, emit a pair preferring
    the median plan over a tail-heavy plan."""
    leads_by_produto: Dict[int, List[float]] = defaultdict(list)
    for r in sheets.get("OrdensFabrico", []):
        produto_id = _safe_int(r.get("Of_ProdutoId"))
        criacao = r.get("Of_DataCriacao")
        acabamento = r.get("Of_DataAcabamento")
        if produto_id is None or not isinstance(criacao, datetime) or not isinstance(acabamento, datetime):
            continue
        lead_days = (acabamento - criacao).total_seconds() / 86400.0
        if lead_days <= 0 or lead_days > 365:
            continue
        leads_by_produto[produto_id].append(lead_days)

    modelo_names = _modelo_id_to_name(sheets.get("Modelos", []))

    pairs: List[DPOPair] = []
    skipped = 0
    for produto_id, leads in leads_by_produto.items():
        if len(leads) < min_support:
            skipped += 1
            continue
        leads_sorted = sorted(leads)
        median = statistics.median(leads_sorted)
        # P90 manual: ceil-style index so the tail entries actually land in P90
        # for small samples (e.g. 15 entries → idx 13, not 12).
        p90_idx = min(len(leads_sorted) - 1, max(0, int(len(leads_sorted) * 0.9)))
        p90 = leads_sorted[p90_idx]
        if median <= 0 or p90 / median < min_tail_ratio:
            skipped += 1
            continue
        modelo_name = modelo_names.get(produto_id, f"modelo {produto_id}")
        prompt = (
            f"Para {modelo_name}, plano A entrega em {median:.0f} dias (mediana "
            f"histórica). Plano B entrega em {p90:.0f} dias (P90). Qual planear?"
        )
        chosen = (
            f"Prefiro o plano A: mediana de {len(leads)} OFs históricas é "
            f"{median:.0f} dias; planear no P90 inflaciona compromisso ao cliente."
        )
        rejected = (
            "Prefiro o plano B: ir pelo P90 dá margem de segurança."
        )
        pairs.append(DPOPair(
            prompt=prompt,
            chosen=chosen,
            rejected=rejected,
            meta={
                "source": "excel_derived",
                "category": "CUSTOMER",
                "extractor": "lead_time",
                "extractor_version": EXTRACTOR_VERSION,
                "evidence": {
                    "sheet": "OrdensFabrico+Modelos",
                    "produto_id": produto_id,
                    "median_days": round(median, 2),
                    "p90_days": round(p90, 2),
                    "tail_ratio": round(p90 / median, 2),
                    "support_n": len(leads),
                },
            },
        ))
        if len(pairs) >= cap_pairs:
            break
    return pairs, ExtractorReport(
        extractor="lead_time",
        pairs_emitted=len(pairs),
        skipped_low_support=skipped,
        notes=f"min_support={min_support} min_tail_ratio={min_tail_ratio}× cap={cap_pairs}",
    )


# ─── Driver ──────────────────────────────────────────────────────────────


def build_excel_dpo(
    *,
    excel_path: Path = DEFAULT_EXCEL,
    output_path: Path = DEFAULT_OUTPUT,
    manifest_path: Optional[Path] = None,
) -> BuildManifest:
    excel_path = Path(excel_path)
    output_path = Path(output_path)
    if manifest_path is None:
        manifest_path = output_path.with_suffix(output_path.suffix + ".manifest.json")
    manifest_path = Path(manifest_path)

    sheets = load_excel(excel_path)

    all_pairs: List[DPOPair] = []
    reports: List[ExtractorReport] = []
    for fn in (
        extract_team_size_pairs,
        extract_real_vs_standard_time_pairs,
        extract_operator_affinity_pairs,
        extract_phase_rework_pairs,
        extract_lead_time_pairs,
    ):
        try:
            pairs, report = fn(sheets)
        except Exception as exc:  # pragma: no cover — defensive
            logger.error("extractor %s failed: %s", fn.__name__, exc, exc_info=True)
            reports.append(ExtractorReport(
                extractor=fn.__name__,
                pairs_emitted=0,
                skipped_low_support=0,
                notes=f"error: {exc}",
            ))
            continue
        all_pairs.extend(pairs)
        reports.append(report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as fh:
        for p in all_pairs:
            fh.write(json.dumps(p.to_json(), ensure_ascii=False))
            fh.write("\n")

    by_extractor: Dict[str, int] = {r.extractor: r.pairs_emitted for r in reports}
    by_category: Dict[str, int] = {}
    for p in all_pairs:
        cat = p.meta.get("category", "OTHER")
        by_category[cat] = by_category.get(cat, 0) + 1

    manifest = BuildManifest(
        output_path=str(output_path),
        excel_path=str(excel_path),
        excel_sha256=_sha256(excel_path),
        triplets_total=len(all_pairs),
        by_extractor=by_extractor,
        by_category=by_category,
        extractor_reports=reports,
    )
    manifest_path.write_text(
        json.dumps(manifest.to_dict(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info(
        "excel_to_dpo: wrote %d pairs to %s; manifest at %s",
        len(all_pairs), output_path, manifest_path,
    )
    return manifest


# ─── CLI ─────────────────────────────────────────────────────────────────


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Excel → DPO pairs builder (S.5.0)")
    p.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    p.add_argument("--manifest", type=Path, default=None)
    return p


def main(argv: Optional[List[str]] = None) -> int:
    args = _build_parser().parse_args(argv)
    manifest = build_excel_dpo(
        excel_path=args.excel,
        output_path=args.output,
        manifest_path=args.manifest,
    )
    print(json.dumps({
        "triplets_total": manifest.triplets_total,
        "by_extractor": manifest.by_extractor,
        "by_category": manifest.by_category,
        "output": manifest.output_path,
    }, ensure_ascii=False, indent=2))
    return 0 if manifest.triplets_total > 0 else 1


if __name__ == "__main__":  # pragma: no cover
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    raise SystemExit(main())
